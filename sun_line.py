import openpyxl
from scipy.stats import norm
import scipy.stats as stats
import numpy as np
from data import *
# 向excel填写最左侧三列
def write_outline(outline):
    file_path = "属性表.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    header_row = 3
    for i, row_data in enumerate(outline):
        sheet.cell(row=i+header_row, column=1, value=row_data[0])
        sheet.cell(row=i+header_row, column=2, value=row_data[1])
        sheet.cell(row=i+header_row, column=3, value=row_data[2])
    wb.save(file_path)

# 获取转职加成
def get_fixed_attr(job, attr_index):
    wb = openpyxl.load_workbook("转职加成.xlsx", data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] == job:
            fixed_attr = [
                row[1+attr_index],
                row[8+attr_index],
                row[15+attr_index],
                row[22+attr_index],
                row[29+attr_index],
            ]
            return fixed_attr

def my_gen_p(mean, varience):
    mindiff = 100000000000
    bestp = [0,0,0,0,0]
    for p0 in range(1, 20):
        for p1 in range(1, 20-p0):
            for p2 in range(1, 20-p0-p1):
                for p3 in range(1, 20-p0-p1-p2):
                    p4 = 20-p0-p1-p2-p3
                    p = [0.05*p0,0.05*p1,0.05*p2,0.05*p3,0.05*p4]
                    m = 0
                    v = 0
                    for i in range(0,5):
                        m += i*p[i]
                    for i in range(0,5):
                        v += p[i]*(i-m)**2
                    if (m-mean)**4 + 2 * (v-varience)**2 < mindiff:
                        bestp = p
                        mindiff = (m-mean)**2
    for i in range(0,5):
        bestp[i] = round(bestp[i],2)
    return bestp

import numpy as np
from scipy.optimize import minimize

def gen_p(expected_value, variance):
    # 定义目标函数，尝试匹配期望和方差
    def objective(p):
        return 0  # 仅寻找可行解，不需要优化目标

    # 约束条件
    def constraint_sum(p):
        return np.sum(p) - 1  # 确保概率和为 1
    
    def constraint_expectation(p):
        values = np.array([0, 1, 2, 3, 4])
        return np.sum(values * p) - expected_value  # 确保期望正确

    def constraint_variance(p):
        values = np.array([0, 1, 2, 3, 4])
        mean = expected_value
        expected_x2 = np.sum((values ** 2) * p)
        return expected_x2 - mean**2 - variance  # 确保方差正确

    # 初始猜测
    initial_guess = np.ones(5) / 5
    
    # 约束条件
    constraints = [
        {"type": "eq", "fun": constraint_sum},
        {"type": "eq", "fun": constraint_expectation},
        {"type": "eq", "fun": constraint_variance},
    ]
    
    # 确保概率在 0 到 1 之间
    bounds = [(0, 1)] * 5
    
    # 求解
    result = minimize(objective, initial_guess, bounds=bounds, constraints=constraints)
    
    if result.success:
        return result.x
    else:
        raise ValueError("无法找到满足要求的概率分布")

# 获取模型倾向
def get_variable_attr(model, attr_index):
    start = 0
    data = []
    data2 = []
    wb = openpyxl.load_workbook("统计表.xlsx", data_only=True)
    ws = wb.active

    # 统计模型初始值
    jobs = []
    cnt = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] == model:
            if row[1] not in jobs:
                jobs.append(row[1])
            cnt += 1
            start += row[2+attr_index]
    start = round(start/cnt, 1)
    
    # 估计模型成长
    for job in jobs:
        fixed_attr = get_fixed_attr(job, attr_index)[0]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if row[0] == model and row[1] == job:
                data.append((row[18 + attr_index] - fixed_attr))
                data2.append((row[18 + attr_index] - fixed_attr) ** 2)
    kurtosis = stats.kurtosis(data, fisher=False)
    M1 = np.mean(data)
    M2 = np.mean(data2)
    M2 = M2 - M1 ** 2
    M1 /= 126
    M2 /= 126
    p = gen_p(M1,M2)
    return [M1, M2, kurtosis, p[0], p[1], p[2], p[3], p[4], round(start)]

# 向excel填写模型倾向和转职加成
def write_data():
    file_path = "属性表.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
        print("正在统计: " + row[2] + " -> " + row[0])
        attr_index = 0 if row[1] == "力" else 1 if row[1] == "魔" else 2 if row[1] == "技" else 3 if row[1] == "速" else 4 if row[1] == "体" else 5
        variable_attr = get_variable_attr(row[2], attr_index)  
        fixed_attr = get_fixed_attr(row[0], attr_index) 
        for j in range(0, len(variable_attr)):
            ws.cell(row=i, column=j+4, value=variable_attr[j])
        for j in range(0, len(fixed_attr)):
            ws.cell(row=i, column=len(variable_attr)+j+4, value=fixed_attr[j])
    wb.save(file_path)

# 向excel填写日10线
def write_sun_line():
    file_path = "属性表.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
        print("计算日10线: " + row[2] + " -> " + row[0])
        s = float(row[col["初始"]])
        m = float(row[col["期望"]]) * 293
        v = float(row[col["方差"]]) * 293
        f = float(row[col["总计"]])
        ws.cell(row=i, column=col["一般"]+1, value=5 * round((s + f + norm.ppf(1-0.1,      loc=m, scale=v**0.5)) / 5))
        ws.cell(row=i, column=col["极品"]+1, value=5 * round((s + f + norm.ppf(1-0.01,     loc=m, scale=v**0.5)) / 5))
        ws.cell(row=i, column=col["大极"]+1, value=5 * round((s + f + norm.ppf(1-0.0001,   loc=m, scale=v**0.5)) / 5))
        ws.cell(row=i, column=col["镇堡"]+1, value=5 * round((s + f + norm.ppf(1-0.000001, loc=m, scale=v**0.5)) / 5))
    wb.save(file_path)

def count_model():
    models = []
    wb = openpyxl.load_workbook("统计表.xlsx", data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] not in models:
            models.append(row[0])
    for model in models:
        cnt = 0
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if row[0] == model:
                cnt += 1
        print(model, " ", cnt)