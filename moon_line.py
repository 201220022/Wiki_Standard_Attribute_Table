
import numpy as np
import openpyxl
from scipy.optimize import root
import random
from data import *
'''
def adjust_rounded_ps(rounded_ps):
    total = sum(rounded_ps)
    diff = round(1.00 - total, 2)
    if diff == 0:
        return rounded_ps
    tolerance = 1e-9
    indices = sorted(range(5), key=lambda i: rounded_ps[i], reverse=True)
    for i in indices:
        current = rounded_ps[i]
        new_val = current + diff
        new_val_rounded = round(new_val, 2)
        if 0 <= new_val_rounded <= 1.0:
            new_ps = rounded_ps.copy()
            new_ps[i] = new_val_rounded
            new_total = sum(new_ps)
            if abs(new_total - 1.00) < tolerance:
                return new_ps
    return rounded_ps
'''

# 返回给定4星5星月百日10线、模型倾向，计算资源的期望消耗量
# 这里的线把转职加成和初始属性先抠了，最后一起加回来
def sum_cost(四星线, 五星线, 月百线, 日十线, p, cost):
    resource_sum = 0

    # 100的位置是0属性对应的位置，这里错开的目的是统一迭代的表达式，加速计算
    dist = np.zeros(905)
    dist[100] = 1
    四星线 = 四星线 + 100
    五星线 = 五星线 + 100
    月百线 = 月百线 + 100
    日十线 = 日十线 + 100

    # 计算到了4星50的分布情况
    for _ in range(126):
        dist[100:905] = (
            p[0] * dist[100:905] +
            p[1] * dist[99:904] +
            p[2] * dist[98:903] +
            p[3] * dist[97:902] +
            p[4] * dist[96:901]
        )
    dist[900] += np.sum(dist[901:905])  #800属性已经镇堡的不能再镇堡了，后面加几已经没所谓了，不影响概率分布的分析。
    dist[901:905] = 0  

    #至此，100%的英雄被升级到四星，资源消耗量为100% * cost[0]
    resource_sum += cost[0]

    #分布完成之后截断达不到四星线的部分
    rest_part = np.sum(dist[四星线:901])
    dist[100:四星线] = 0

    # 计算到了5星60的分布情况
    for _ in range(59):
        dist[100:905] = (
            p[0] * dist[100:905] +
            p[1] * dist[99:904] +
            p[2] * dist[98:903] +
            p[3] * dist[97:902] +
            p[4] * dist[96:901]
        )
    dist[900] += np.sum(dist[901:905])
    dist[901:905] = 0  

    #至此，rest_part的英雄被升级到五星，资源消耗量为rest_part * cost[1]
    resource_sum += rest_part * cost[1]

    #分布完成之后截断达不到五星线的部分
    rest_part = np.sum(dist[五星线:901])
    dist[四星线:五星线] = 0

    # 计算到了月百的分布情况
    for _ in range(99):
        dist[100:905] = (
            p[0] * dist[100:905] +
            p[1] * dist[99:904] +
            p[2] * dist[98:903] +
            p[3] * dist[97:902] +
            p[4] * dist[96:901]
        )
    dist[900] += np.sum(dist[901:905])
    dist[901:905] = 0  

    #至此，rest_part的英雄被升级到月百，资源消耗量为rest_part * cost[2]
    resource_sum += rest_part * cost[2]

    #分布完成之后截断达不到月百线的部分
    rest_part = np.sum(dist[月百线:901])
    dist[五星线:月百线] = 0

    # 计算到了日十的分布情况
    for _ in range(9):
        dist[100:905] = (
            p[0] * dist[100:905] +
            p[1] * dist[99:904] +
            p[2] * dist[98:903] +
            p[3] * dist[97:902] +
            p[4] * dist[96:901]
        )
    dist[900] += np.sum(dist[901:905])
    dist[901:905] = 0  

    #至此，rest_part的英雄被升级到日十，资源消耗量为rest_part * cost[3]
    resource_sum += rest_part * cost[3]

    #分布完成之后截断达不到日十线的部分，也就是说，计算超过了日10线的占比 -> 出货率
    rest_part = np.sum(dist[日十线:901])
    
    #升出一个过线冒险者的资源消耗期望，等于刚才这一轮分布的资源消耗，除以出货率。
    resource_sum /= rest_part
    return round(resource_sum)

def optimize_lines(t, p, cost):
    """
    在 0 <= l1 <= l2 <= l3 <= t 范围内，寻找使 sum_cost(l1, l2, l3, t, m, v) 最小的整数组合。
    这里先根据比例（126/293, 185/293, 284/293）给出初始估计，然后采用多步长局部搜索优化。
    """
    # 初始估计
    l1 = round(126 * t / 293)
    l2 = round(185 * t / 293)
    l3 = round(284 * t / 293)
    # 保证满足约束
    l1 = max(0, min(l1, t))
    l2 = max(l1, min(l2, t))
    l3 = max(l2, min(l3, t))
    
    best_cost = sum_cost(l1, l2, l3, t, p, cost)
    step = 8  # 初始步长，可根据 t 调整
    
    while step > 0:
        improved = True
        while improved:
            improved = False
            # 尝试调整 l1
            for delta in [-step, step]:
                new_l1 = l1 + delta
                if 0 <= new_l1 <= l2:
                    c = sum_cost(new_l1, l2, l3, t, p, cost)
                    if c < best_cost:
                        best_cost = c
                        l1 = new_l1
                        improved = True
                        break
            if improved:
                continue
            # 尝试调整 l2
            for delta in [-step, step]:
                new_l2 = l2 + delta
                if l1 <= new_l2 <= l3:
                    c = sum_cost(l1, new_l2, l3, t, p, cost)
                    if c < best_cost:
                        best_cost = c
                        l2 = new_l2
                        improved = True
                        break
            if improved:
                continue
            # 尝试调整 l3
            # for delta in [-step, step]:
            #     new_l3 = l3 + delta
            #     if l2 <= new_l3 <= t:
            #         c = sum_cost(l1, l2, new_l3, t, p, cost)
            #         if c < best_cost:
            #             best_cost = c
            #             l3 = new_l3
            #             improved = True
            #             break
        step //= 2  # 缩小步长
    return l1, l2, l3

# 向excel填写月百五星四星线
def write_moon_line():
    file_path = "属性表.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
        print("计算月百五星四星线: " + row[2] + " -> " + row[0])
        s = int(row[col["初始"]])
        f = int(row[col["总计"]])
        a4 = int(row[col["四星"]])
        a5 = int(row[col["五星"]])
        am = int(row[col["月百"]])
        p0 = float(row[col["+0"]])
        p1 = float(row[col["+1"]])
        p2 = float(row[col["+2"]])
        p3 = float(row[col["+3"]])
        p4 = float(row[col["+4"]])
        p = [p0,p1,p2,p3,p4]
        
        for typ in ["一般", "极品", "大极", "镇堡"]:
            # 日十减去初始和转职部分
            t = row[col[typ]] - s - f

            # 分情况应用cost
            cost = [52354, 16206, 226880, 928400]
            if typ == "一般":
                cost[0] = 85144
                
            # 计算最合适的line取值
            line1, line2, line3 = optimize_lines(t, p, cost)

            line1 += s + a4
            line2 += s + a4 + a5
            line3 += s + a4 + a5 + am
            line1 = 5 * round(line1/5)
            line2 = 5 * round(line2/5)
            # line3 = 5 * round(line3/5)
            print(line1,line2,line3,row[col[typ]])
            cost = sum_cost(line1 - s - a4, line2 - s - a4 - a5, line3 - s - a4 - a5 - am, t, p, cost)
            ws.cell(row=i, column=col[typ], value=line3)
            ws.cell(row=i, column=col[typ]-1, value=line2)
            ws.cell(row=i, column=col[typ]-2, value=line1)
            ws.cell(row=i, column=col[typ]+2, value=round(cost/10000))

    wb.save(file_path)


'''
def optimize_lines(t, p, cost):
    # 初始估计
    l1 = round(126 * t / 293)
    l2 = round(185 * t / 293)
    l3 = round(284 * t / 293)
    best_cost = sum_cost(l1, l2, l3, t, p, cost)
    while True:
        bi = 0
        bj = 0
        bk = 0
        bc = best_cost
        for i in range(-2,3):
            for j in range(-2,3):
                for k in range(-1,2):
                    c = sum_cost(l1+i, l2+j, l3+k, t, p, cost)
                    if c < bc:
                        bc = c
                        bi = i
                        bj = j
                        bk = k
        if bc == best_cost:
            break
        l1 += bi
        l2 += bj
        l3 += bk
        best_cost = bc
    return l1, l2, l3
'''